"""Unit tests for Spec 01I (Onboarding flow) — backend pieces.

Covers the parts that don't require a live Postgres:
- Pricing parser (xlsx): valid file, invalid rows, missing columns
- Pricing template generator: returns a real xlsx
- Onboarding state machine: forward-only transitions, completion stamp,
  banner show/hide logic, dismiss timestamp
- POST /v1/jobs/batch validation: rejects > 10, rejects empty, rejects
  empty address; UI label -> enum mapping
- POST /v1/company: full profile pass-through to rpc_onboard_user
- GET /v1/company/onboarding-status response shape

Tests that require real Postgres (advisory locks, CHECK constraints,
RLS policies, RPC atomicity) live in tests/integration/test_onboarding_integration.py
and run only when local Supabase is up.
"""

from __future__ import annotations

import io
from unittest.mock import MagicMock, patch
from uuid import uuid4

import pytest
from openpyxl import Workbook, load_workbook

from api.auth.schemas import (
    ONBOARDING_STEP_ORDER,
    OnboardingStatusResponse,
)
from api.config import settings
from api.pricing.service import (
    build_template_xlsx,
    errors_to_csv,
    get_error_report,
    parse_pricing_xlsx,
)
from tests.conftest import AsyncSupabaseMock, make_mock_supabase

# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------


def _make_xlsx_bytes(rows: list[tuple], *, sheet_title: str = "Tier A") -> bytes:
    """Build an xlsx in-memory from a list of row tuples."""
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_title
    for r in rows:
        ws.append(list(r))
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _user_row(
    user_id,
    company_id,
    *,
    auth_user_id=None,
    step="company_profile",
    completed_at=None,
    dismissed_at=None,
):
    """Build a fake users-table row for service tests.

    Includes ``auth_user_id`` distinct from ``id`` because the production
    code looks up by ``auth_user_id`` for status reads / state writes (the
    Supabase JWT carries auth_user_id in ``sub``, not users.id). Earlier
    fixtures omitted this field, so a service bug that confused the two
    keys was invisible to the test mocks. Caught by code review.
    """
    return {
        "id": str(user_id),
        "auth_user_id": str(auth_user_id or user_id),
        "company_id": str(company_id),
        "onboarding_step": step,
        "onboarding_completed_at": completed_at,
        "setup_banner_dismissed_at": dismissed_at,
        "role": "owner",
        "is_platform_admin": False,
    }


# ---------------------------------------------------------------------------
# 1. Pricing template
# ---------------------------------------------------------------------------


class TestPricingTemplate:
    """build_template_xlsx() returns a real .xlsx with the expected shape."""

    def test_template_returns_valid_xlsx(self):
        body = build_template_xlsx()
        # Smoke check: openpyxl can load it back
        wb = load_workbook(io.BytesIO(body), read_only=True)
        assert wb.sheetnames == ["Tier A"]
        ws = wb["Tier A"]
        rows = list(ws.iter_rows(values_only=True))
        # Header + 3 sample rows
        assert len(rows) == 4
        assert tuple(rows[0]) == ("code", "description", "unit", "price")
        # Sample data plausible
        assert rows[1][0] == "WTR DRYOUT"
        assert isinstance(rows[1][3], (int, float))


# ---------------------------------------------------------------------------
# 2. Pricing parser
# ---------------------------------------------------------------------------


class TestPricingParser:
    """parse_pricing_xlsx — happy path + per-row errors."""

    def test_valid_file_returns_rows_no_errors(self):
        body = _make_xlsx_bytes(
            [
                ("code", "description", "unit", "price"),
                ("WTR DRYOUT", "Water dryout", "SF", 1.25),
                ("DRYWLL RR", "Drywall replace", "SF", 3.50),
            ]
        )
        rows, errors = parse_pricing_xlsx(body)
        assert errors == []
        assert len(rows) == 2
        assert rows[0] == {
            "code": "WTR DRYOUT",
            "description": "Water dryout",
            "unit": "SF",
            "price": 1.25,
            "tier": "A",
        }

    def test_missing_required_column_returns_error(self):
        body = _make_xlsx_bytes(
            [
                ("description", "unit", "price"),  # no 'code'
                ("Water dryout", "SF", 1.25),
            ]
        )
        rows, errors = parse_pricing_xlsx(body)
        assert rows == []
        assert any(e.field == "code" and "missing" in e.message.lower() for e in errors)

    def test_blank_code_returns_row_error(self):
        body = _make_xlsx_bytes(
            [
                ("code", "description", "unit", "price"),
                ("", "Water dryout", "SF", 1.25),
                ("DRYWLL RR", "Drywall", "SF", 3.50),
            ]
        )
        rows, errors = parse_pricing_xlsx(body)
        assert rows == []
        assert any(e.row == 2 and e.field == "code" for e in errors)

    def test_blank_price_returns_row_error(self):
        body = _make_xlsx_bytes(
            [
                ("code", "price"),
                ("WTR DRYOUT", None),
            ]
        )
        rows, errors = parse_pricing_xlsx(body)
        assert rows == []
        assert any(e.row == 2 and e.field == "price" for e in errors)

    def test_negative_price_returns_row_error(self):
        body = _make_xlsx_bytes(
            [
                ("code", "price"),
                ("WTR DRYOUT", -5),
            ]
        )
        rows, errors = parse_pricing_xlsx(body)
        assert rows == []
        assert any(e.row == 2 and e.field == "price" and ">= 0" in e.message for e in errors)

    def test_currency_string_price_accepted(self):
        body = _make_xlsx_bytes(
            [
                ("code", "price"),
                ("WTR DRYOUT", "$1,250.00"),
            ]
        )
        rows, errors = parse_pricing_xlsx(body)
        assert errors == []
        assert rows[0]["price"] == 1250.00

    def test_unparseable_price_string_returns_row_error(self):
        body = _make_xlsx_bytes(
            [
                ("code", "price"),
                ("WTR DRYOUT", "not-a-number"),
            ]
        )
        rows, errors = parse_pricing_xlsx(body)
        assert rows == []
        assert any(e.row == 2 and e.field == "price" for e in errors)

    def test_blank_rows_skipped_silently(self):
        body = _make_xlsx_bytes(
            [
                ("code", "description", "unit", "price"),
                ("WTR DRYOUT", "Water dryout", "SF", 1.25),
                (None, None, None, None),  # blank row
                ("DRYWLL RR", "Drywall", "SF", 3.50),
            ]
        )
        rows, errors = parse_pricing_xlsx(body)
        assert errors == []
        assert len(rows) == 2

    def test_duplicate_code_in_file_returns_error(self):
        body = _make_xlsx_bytes(
            [
                ("code", "price"),
                ("WTR DRYOUT", 1.25),
                ("WTR DRYOUT", 1.50),
            ]
        )
        rows, errors = parse_pricing_xlsx(body)
        assert rows == []
        assert any(e.row == 3 and "duplicate" in e.message.lower() for e in errors)

    def test_tier_extracted_from_sheet_title(self):
        body = _make_xlsx_bytes(
            [("code", "price"), ("WTR DRYOUT", 1.25)],
            sheet_title="Tier B",
        )
        rows, _ = parse_pricing_xlsx(body)
        assert rows[0]["tier"] == "B"

    def test_malformed_workbook_returns_single_error(self):
        rows, errors = parse_pricing_xlsx(b"this is not a real xlsx file")
        assert rows == []
        assert len(errors) == 1
        assert "cannot read" in errors[0].message.lower()

    def test_errors_to_csv_round_trip(self):
        body = _make_xlsx_bytes(
            [
                ("code", "price"),
                ("", 1.25),
            ]
        )
        _, errors = parse_pricing_xlsx(body)
        csv_text = errors_to_csv(errors)
        # Header + at least one row
        lines = csv_text.strip().split("\n")
        assert lines[0] == "row,field,message"
        assert len(lines) >= 2

    def test_csv_escapes_quotes_and_commas(self):
        from api.pricing.schemas import PricingRowError

        errors = [
            PricingRowError(row=2, field="description", message='Has, comma "and quotes"'),
        ]
        csv_text = errors_to_csv(errors)
        # The whole field should be wrapped in quotes, with internal quotes doubled
        assert '"Has, comma ""and quotes"""' in csv_text


# ---------------------------------------------------------------------------
# 3. Pricing error report storage
# ---------------------------------------------------------------------------


class TestPricingErrorReport:
    def test_report_stored_and_retrieved(self):
        from api.pricing.schemas import PricingRowError
        from api.pricing.service import _store_error_report

        company_id = uuid4()
        errors = [PricingRowError(row=2, field="code", message="bad")]
        run_id = _store_error_report(company_id, errors)

        retrieved = get_error_report(run_id, company_id=company_id)
        assert retrieved is not None
        assert retrieved[0].field == "code"

    def test_report_not_found_returns_none(self):
        assert get_error_report("nonexistent-run-id", company_id=uuid4()) is None

    def test_report_cross_tenant_lookup_returns_none(self):
        """Tenant A cannot fetch tenant B's report by guessing a run_id."""
        from api.pricing.schemas import PricingRowError
        from api.pricing.service import _store_error_report

        company_a = uuid4()
        company_b = uuid4()
        errors = [PricingRowError(row=2, field="code", message="leaked")]
        run_id = _store_error_report(company_a, errors)

        assert get_error_report(run_id, company_id=company_b) is None
        # Sanity: company_a still gets it back
        assert get_error_report(run_id, company_id=company_a) is not None


# ---------------------------------------------------------------------------
# 4. Onboarding state machine
# ---------------------------------------------------------------------------


class TestOnboardingStateMachine:
    """get_onboarding_status / update_onboarding_step / dismiss_setup_banner.

    All three exercise the same in-memory mock so we can verify the
    derived ``has_jobs``/``has_pricing``/``show_setup_banner`` logic.
    """

    @pytest.fixture
    def user_id(self):
        return uuid4()

    @pytest.fixture
    def company_id(self):
        return uuid4()

    def _patch_admin_client(self, client):
        return patch(
            "api.auth.service.get_supabase_admin_client",
            return_value=client,
        )

    def _build_status_mock(
        self,
        user_row,
        *,
        has_jobs: bool,
        has_pricing: bool,
        update_returns_user: bool = True,
    ):
        """Return an AsyncSupabaseMock configured for get_onboarding_status."""
        mock_client = AsyncSupabaseMock()

        def table_side_effect(table_name):
            t = AsyncSupabaseMock()
            if table_name == "users":
                # get_onboarding_status: select.eq.is_.maybe_single.execute.data = user_row
                (
                    t.select.return_value.eq.return_value.is_.return_value.maybe_single.return_value.execute.return_value
                ).data = user_row
                # update_onboarding_step / dismiss_setup_banner: update.eq.execute / update.eq.is_.execute
                if update_returns_user:
                    t.update.return_value.eq.return_value.execute.return_value.data = [user_row]
                    (
                        t.update.return_value.eq.return_value.is_.return_value.execute.return_value
                    ).data = [user_row]
                else:
                    t.update.return_value.eq.return_value.execute.return_value.data = []
                    (
                        t.update.return_value.eq.return_value.is_.return_value.execute.return_value
                    ).data = []
            elif table_name == "jobs":
                result = MagicMock()
                result.data = [{"id": str(uuid4())}] if has_jobs else []
                # _exists path: select.eq.limit.is_.execute / select.eq.is_.limit.execute
                # (we use eq.limit().is_().execute() in code: client.table.select.eq.limit then is_)
                (
                    t.select.return_value.eq.return_value.limit.return_value.is_.return_value.execute.return_value
                ) = result
                (
                    t.select.return_value.eq.return_value.is_.return_value.limit.return_value.execute.return_value
                ) = result
            elif table_name == "scope_codes":
                result = MagicMock()
                result.data = [{"id": str(uuid4())}] if has_pricing else []
                # scope_codes has no deleted_at filter — code path is
                # select.eq.limit.execute (no is_)
                (
                    t.select.return_value.eq.return_value.limit.return_value.execute.return_value
                ) = result
            return t

        mock_client.table.side_effect = table_side_effect
        return mock_client

    @pytest.mark.asyncio
    async def test_get_onboarding_status_derives_has_jobs(self, user_id, company_id):
        from api.auth.service import get_onboarding_status

        user_row = _user_row(user_id, company_id)
        mock_client = self._build_status_mock(user_row, has_jobs=True, has_pricing=False)

        with self._patch_admin_client(mock_client):
            status = await get_onboarding_status(user_id)

        assert isinstance(status, OnboardingStatusResponse)
        assert status.has_jobs is True
        assert status.has_pricing is False
        assert status.has_company is True
        assert status.step == "company_profile"

    @pytest.mark.asyncio
    async def test_show_setup_banner_only_when_completed_undismissed_no_pricing(
        self, user_id, company_id
    ):
        from api.auth.service import get_onboarding_status

        user_row = _user_row(
            user_id,
            company_id,
            step="complete",
            completed_at="2026-04-27T00:00:00Z",
            dismissed_at=None,
        )
        mock_client = self._build_status_mock(user_row, has_jobs=True, has_pricing=False)
        with self._patch_admin_client(mock_client):
            status = await get_onboarding_status(user_id)

        assert status.show_setup_banner is True

    @pytest.mark.asyncio
    async def test_setup_banner_hidden_when_dismissed(self, user_id, company_id):
        from api.auth.service import get_onboarding_status

        user_row = _user_row(
            user_id,
            company_id,
            step="complete",
            completed_at="2026-04-27T00:00:00Z",
            dismissed_at="2026-04-27T01:00:00Z",
        )
        mock_client = self._build_status_mock(user_row, has_jobs=False, has_pricing=False)
        with self._patch_admin_client(mock_client):
            status = await get_onboarding_status(user_id)

        assert status.show_setup_banner is False

    @pytest.mark.asyncio
    async def test_setup_banner_hidden_when_pricing_uploaded(self, user_id, company_id):
        from api.auth.service import get_onboarding_status

        user_row = _user_row(
            user_id,
            company_id,
            step="complete",
            completed_at="2026-04-27T00:00:00Z",
            dismissed_at=None,
        )
        mock_client = self._build_status_mock(user_row, has_jobs=True, has_pricing=True)
        with self._patch_admin_client(mock_client):
            status = await get_onboarding_status(user_id)

        assert status.show_setup_banner is False

    @pytest.mark.asyncio
    async def test_setup_banner_hidden_before_complete(self, user_id, company_id):
        """Banner only shows after onboarding completes (Decision Log #5)."""
        from api.auth.service import get_onboarding_status

        user_row = _user_row(user_id, company_id, step="pricing", completed_at=None)
        mock_client = self._build_status_mock(user_row, has_jobs=True, has_pricing=False)
        with self._patch_admin_client(mock_client):
            status = await get_onboarding_status(user_id)

        assert status.show_setup_banner is False

    @pytest.mark.asyncio
    async def test_update_onboarding_step_forward_succeeds(self, user_id, company_id):
        from api.auth.service import update_onboarding_step

        user_row = _user_row(user_id, company_id, step="company_profile")
        # After update, return user with new step so the follow-up
        # get_onboarding_status sees the new value.
        updated = dict(user_row, onboarding_step="jobs_import")
        mock_client = self._build_status_mock(updated, has_jobs=False, has_pricing=False)
        # Serve user_row for the initial select, updated for the post-update select
        # (the simple mock structure returns the same row from both reads — that
        # matches how the real DB would behave after the update commits).
        with self._patch_admin_client(mock_client):
            status = await update_onboarding_step(user_id, "jobs_import")

        assert status.step == "jobs_import"

    @pytest.mark.asyncio
    async def test_update_onboarding_step_backward_rejected(self, user_id, company_id):
        from api.auth.service import update_onboarding_step
        from api.shared.exceptions import AppException

        user_row = _user_row(user_id, company_id, step="pricing")
        mock_client = self._build_status_mock(user_row, has_jobs=False, has_pricing=False)

        with self._patch_admin_client(mock_client), pytest.raises(AppException) as exc_info:
            await update_onboarding_step(user_id, "company_profile")

        assert exc_info.value.status_code == 400
        assert exc_info.value.error_code == "ONBOARDING_BACKWARD_TRANSITION"

    @pytest.mark.asyncio
    async def test_update_onboarding_step_invalid_value_rejected(self, user_id, company_id):
        from api.auth.service import update_onboarding_step
        from api.shared.exceptions import AppException

        with pytest.raises(AppException) as exc_info:
            await update_onboarding_step(user_id, "not-a-real-step")

        assert exc_info.value.error_code == "INVALID_ONBOARDING_STEP"

    @pytest.mark.asyncio
    async def test_update_to_complete_stamps_completed_at(self, user_id, company_id):
        from api.auth.service import update_onboarding_step

        # Capture the update payload to assert it carries onboarding_completed_at
        captured: dict = {}

        mock_client = AsyncSupabaseMock()

        def table_side_effect(table_name):
            t = AsyncSupabaseMock()
            if table_name == "users":
                # Initial select returns user at first_job step
                (
                    t.select.return_value.eq.return_value.is_.return_value.maybe_single.return_value.execute.return_value
                ).data = _user_row(user_id, company_id, step="first_job")

                # Capture the update kwargs
                def update_capture(payload):
                    captured["payload"] = payload
                    next_t = AsyncSupabaseMock()
                    next_t.eq.return_value.execute.return_value.data = [
                        _user_row(
                            user_id,
                            company_id,
                            step="complete",
                            completed_at="2026-04-27T00:00:00Z",
                        )
                    ]
                    return next_t

                t.update.side_effect = update_capture
            elif table_name == "jobs":
                r = MagicMock()
                r.data = []
                (
                    t.select.return_value.eq.return_value.limit.return_value.is_.return_value.execute.return_value
                ) = r
                (
                    t.select.return_value.eq.return_value.is_.return_value.limit.return_value.execute.return_value
                ) = r
            elif table_name == "scope_codes":
                r = MagicMock()
                r.data = []
                (t.select.return_value.eq.return_value.limit.return_value.execute.return_value) = r
            return t

        mock_client.table.side_effect = table_side_effect

        with patch(
            "api.auth.service.get_supabase_admin_client",
            return_value=mock_client,
        ):
            await update_onboarding_step(user_id, "complete")

        assert captured["payload"]["onboarding_step"] == "complete"
        assert "onboarding_completed_at" in captured["payload"]

    @pytest.mark.asyncio
    async def test_dismiss_setup_banner_sets_timestamp(self, user_id, company_id):
        from api.auth.service import dismiss_setup_banner

        captured: dict = {}
        mock_client = AsyncSupabaseMock()

        def table_side_effect(table_name):
            t = AsyncSupabaseMock()
            if table_name == "users":
                # Capture the update payload — banner dismissal sets the
                # timestamp; verify exactly that field is stamped.
                def update_capture(payload):
                    captured["payload"] = payload
                    next_t = AsyncSupabaseMock()
                    # update.eq.is_.execute path
                    next_t.eq.return_value.is_.return_value.execute.return_value.data = [
                        _user_row(
                            user_id,
                            company_id,
                            step="complete",
                            completed_at="2026-04-27T00:00:00Z",
                            dismissed_at="2026-04-27T01:00:00Z",
                        )
                    ]
                    # Also satisfy update.eq.execute (older path) just in case
                    next_t.eq.return_value.execute.return_value.data = [
                        _user_row(user_id, company_id, dismissed_at="2026-04-27T01:00:00Z")
                    ]
                    return next_t

                t.update.side_effect = update_capture
                # The follow-up get_onboarding_status select
                (
                    t.select.return_value.eq.return_value.is_.return_value.maybe_single.return_value.execute.return_value
                ).data = _user_row(
                    user_id,
                    company_id,
                    step="complete",
                    completed_at="2026-04-27T00:00:00Z",
                    dismissed_at="2026-04-27T01:00:00Z",
                )
            elif table_name in ("jobs", "scope_codes"):
                r = MagicMock()
                r.data = []
                (
                    t.select.return_value.eq.return_value.limit.return_value.is_.return_value.execute.return_value
                ) = r
                (
                    t.select.return_value.eq.return_value.is_.return_value.limit.return_value.execute.return_value
                ) = r
                (t.select.return_value.eq.return_value.limit.return_value.execute.return_value) = r
            return t

        mock_client.table.side_effect = table_side_effect

        with patch(
            "api.auth.service.get_supabase_admin_client",
            return_value=mock_client,
        ):
            await dismiss_setup_banner(user_id)

        assert "setup_banner_dismissed_at" in captured["payload"]


# ---------------------------------------------------------------------------
# 5. POST /v1/jobs/batch validation
# ---------------------------------------------------------------------------


class TestJobsBatchValidation:
    """Validation logic that doesn't require the RPC to round-trip."""

    @pytest.fixture
    def headers(self, valid_token, jwt_secret):
        return {"Authorization": f"Bearer {valid_token}"}, jwt_secret

    def test_batch_rejects_more_than_10(self, client, mock_user_row, headers):
        auth_headers, jwt_secret = headers
        mock_client = make_mock_supabase(user_row=mock_user_row)

        # 11 minimal jobs — pydantic rejects at the schema layer (max_length=10)
        jobs = [{"address_line1": f"{i} Main St"} for i in range(11)]

        with (
            patch.object(settings, "supabase_jwt_secret", jwt_secret),
            patch("api.auth.middleware.get_supabase_admin_client", return_value=mock_client),
        ):
            resp = client.post("/v1/jobs/batch", json={"jobs": jobs}, headers=auth_headers)

        assert resp.status_code == 422  # pydantic validation

    def test_batch_rejects_empty(self, client, mock_user_row, headers):
        auth_headers, jwt_secret = headers
        mock_client = make_mock_supabase(user_row=mock_user_row)

        with (
            patch.object(settings, "supabase_jwt_secret", jwt_secret),
            patch("api.auth.middleware.get_supabase_admin_client", return_value=mock_client),
        ):
            resp = client.post("/v1/jobs/batch", json={"jobs": []}, headers=auth_headers)

        assert resp.status_code == 422

    def test_batch_rejects_blank_address(self, client, mock_user_row, headers):
        auth_headers, jwt_secret = headers
        mock_client = make_mock_supabase(user_row=mock_user_row)

        with (
            patch.object(settings, "supabase_jwt_secret", jwt_secret),
            patch("api.auth.middleware.get_supabase_admin_client", return_value=mock_client),
        ):
            resp = client.post(
                "/v1/jobs/batch",
                json={"jobs": [{"address_line1": ""}]},
                headers=auth_headers,
            )

        assert resp.status_code == 422

    @pytest.mark.asyncio
    async def test_status_normalize_ui_labels(self):
        """Spec 01K labels (Lead/Active/Invoiced) map to lifecycle enum values."""
        from api.jobs.service import _normalize_batch_status

        assert _normalize_batch_status("Lead") == "lead"
        assert _normalize_batch_status("active") == "active"
        assert _normalize_batch_status("Invoiced") == "invoiced"
        assert _normalize_batch_status(None) == "lead"
        # Pass-through enum values
        assert _normalize_batch_status("lead") == "lead"
        assert _normalize_batch_status("active") == "active"

    @pytest.mark.asyncio
    async def test_status_normalize_rejects_unknown(self):
        from api.jobs.service import _normalize_batch_status
        from api.shared.exceptions import AppException

        with pytest.raises(AppException) as exc_info:
            _normalize_batch_status("not-a-real-status")
        assert exc_info.value.error_code == "INVALID_STATUS"

    def test_batch_invokes_rpc_with_correct_payload(
        self, client, mock_user_row, headers, mock_company_id, mock_user_id
    ):
        """3-job batch maps UI labels and reaches rpc_create_jobs_batch."""
        auth_headers, jwt_secret = headers

        # Set the auth context user/company ids to match the route resolution
        mock_user_row["id"] = str(mock_user_id)
        mock_user_row["company_id"] = str(mock_company_id)

        mock_client = make_mock_supabase(user_row=mock_user_row)

        # Capture RPC arguments
        captured: dict = {}

        def rpc_side_effect(name, params):
            captured["name"] = name
            captured["params"] = params
            inner = AsyncSupabaseMock()
            inner.execute.return_value = MagicMock(
                data={
                    "created": 3,
                    "jobs": [
                        {"job_id": str(uuid4()), "job_number": f"JOB-20260427-{i:03d}"}
                        for i in range(1, 4)
                    ],
                }
            )
            return inner

        mock_admin = AsyncSupabaseMock()
        mock_admin.rpc.side_effect = rpc_side_effect

        body = {
            "jobs": [
                {
                    "address_line1": "123 Main St",
                    "city": "Troy",
                    "state": "MI",
                    "zip": "48083",
                    "loss_type": "water",
                    "status": "Lead",  # Spec 01K — UI label "Lead" → enum "lead"
                },
                {
                    "address_line1": "456 Oak Ave",
                    "city": "Detroit",
                    "state": "MI",
                    "zip": "48201",
                    "loss_type": "fire",
                    "status": "Active",
                },
                {
                    "address_line1": "789 Birch Ln",
                    "loss_type": "mold",
                    "status": "Invoiced",
                },
            ]
        }

        with (
            patch.object(settings, "supabase_jwt_secret", jwt_secret),
            patch("api.auth.middleware.get_supabase_admin_client", return_value=mock_client),
            patch("api.jobs.service.get_supabase_admin_client", return_value=mock_admin),
        ):
            resp = client.post("/v1/jobs/batch", json=body, headers=auth_headers)

        assert resp.status_code == 201, resp.text
        assert resp.json()["created"] == 3
        assert captured["name"] == "rpc_create_jobs_batch"
        # Status labels should be normalized to enum values before hitting RPC
        statuses = [j["status"] for j in captured["params"]["p_jobs"]]
        assert statuses == ["lead", "active", "invoiced"]


# ---------------------------------------------------------------------------
# 6. POST /v1/company — extended profile passes through to RPC
# ---------------------------------------------------------------------------


class TestCompanyCreateExtendedProfile:
    """The new address + service_area fields reach rpc_onboard_user."""

    def test_post_company_passes_full_profile_to_rpc(
        self, client, valid_token, jwt_secret, mock_auth_user_id, mock_company_id, mock_user_id
    ):
        """POST /v1/company with full profile -> RPC called with all fields."""
        # Capture the RPC call args
        captured: dict = {}

        mock_client = AsyncSupabaseMock()
        mock_auth_response = AsyncSupabaseMock()
        mock_auth_response.user.email = "owner@dryco.com"
        mock_auth_response.user.user_metadata = {"full_name": "Owner Name"}
        mock_client.auth.admin.get_user_by_id.return_value = mock_auth_response

        # Spec 01K: onboarding now calls rpc_seed_closeout_settings after
        # rpc_onboard_user. We need to capture each RPC by name so the test
        # below can assert on the rpc_onboard_user call specifically.
        captured_by_name: dict = {}

        def rpc_side_effect(name, params):
            captured_by_name[name] = params
            captured["name"] = name
            captured["params"] = params
            inner = AsyncSupabaseMock()
            if name == "rpc_seed_closeout_settings":
                # Best-effort post-onboarding seed — return empty success.
                inner.execute.return_value = MagicMock(data=None)
                return inner
            company_row = {
                "id": str(mock_company_id),
                "name": params["p_company_name"],
                "slug": params["p_company_slug"],
                "phone": params["p_company_phone"],
                "email": params["p_email"],
                "logo_url": None,
                "address": params["p_company_address"],
                "city": params["p_company_city"],
                "state": params["p_company_state"],
                "zip": params["p_company_zip"],
                "service_area": params["p_service_area"],
                "subscription_tier": "free",
                "created_at": "2026-04-27T00:00:00Z",
                "updated_at": "2026-04-27T00:00:00Z",
            }
            user_row = {
                "id": str(mock_user_id),
                "auth_user_id": str(mock_auth_user_id),
                "company_id": str(mock_company_id),
                "email": params["p_email"],
                "name": params["p_name"],
                "first_name": params["p_first_name"],
                "last_name": params["p_last_name"],
                "phone": None,
                "avatar_url": None,
                "title": None,
                "role": "owner",
                "is_platform_admin": False,
                "deleted_at": None,
            }
            inner.execute.return_value = MagicMock(
                data={"already_exists": False, "user": user_row, "company": company_row}
            )
            return inner

        mock_client.rpc.side_effect = rpc_side_effect

        with (
            patch.object(settings, "supabase_jwt_secret", jwt_secret),
            patch("api.auth.service.get_supabase_admin_client", return_value=mock_client),
            patch("api.shared.database.get_supabase_admin_client", return_value=mock_client),
        ):
            resp = client.post(
                "/v1/company",
                json={
                    "name": "Dryco LLC",
                    "phone": "(586) 555-1212",
                    "address": "100 Main St",
                    "city": "Warren",
                    "state": "MI",
                    "zip": "48089",
                    "service_area": ["Warren/Macomb", "Oakland"],
                },
                headers={"Authorization": f"Bearer {valid_token}"},
            )

        assert resp.status_code == 201, resp.text
        # Assert on the rpc_onboard_user payload (not the last RPC call,
        # which is now rpc_seed_closeout_settings — Spec 01K addition).
        params = captured_by_name["rpc_onboard_user"]
        assert params["p_company_name"] == "Dryco LLC"
        assert params["p_company_phone"] == "(586) 555-1212"
        assert params["p_company_address"] == "100 Main St"
        assert params["p_company_city"] == "Warren"
        assert params["p_company_state"] == "MI"
        assert params["p_company_zip"] == "48089"
        assert params["p_service_area"] == ["Warren/Macomb", "Oakland"]
        # And confirm the closeout-settings seed actually fired for the new company.
        assert "rpc_seed_closeout_settings" in captured_by_name
        assert captured_by_name["rpc_seed_closeout_settings"]["p_company_id"] == str(mock_company_id)


# ---------------------------------------------------------------------------
# 7. ONBOARDING_STEP_ORDER invariants — the contract the state machine relies on
# ---------------------------------------------------------------------------


def test_onboarding_step_order_is_canonical():
    """Spec 01I: order is fixed and 'complete' is last."""
    assert ONBOARDING_STEP_ORDER == (
        "company_profile",
        "jobs_import",
        "pricing",
        "first_job",
        "complete",
    )
    assert ONBOARDING_STEP_ORDER[-1] == "complete"


# ---------------------------------------------------------------------------
# Note on advisory-lock / CHECK-constraint / RPC-atomicity tests
# ---------------------------------------------------------------------------
# These need a real Postgres to exercise (advisory locks aren't a Python
# concept and CHECK constraints fire at INSERT time inside the database).
# They live in tests/integration/test_onboarding_integration.py and run only
# when local Supabase is reachable — see tests/integration/conftest.py.
