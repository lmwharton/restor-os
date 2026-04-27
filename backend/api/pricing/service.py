"""Pricing upload + template service (Spec 01I Phase 3).

Parses the contractor's Xactimate-style price list (.xlsx), validates each
row, and inserts everything in a single transaction. On any validation
failure the whole upload aborts and we return per-row errors keyed by
spreadsheet row number.

Why a single insert + atomic abort? The wizard treats pricing as one
deliverable. Half-loaded prices would silently produce wrong estimates
later — it's safer to make the user fix the file and re-upload. supabase-py
lacks a transaction API, so we use a single ``upsert`` (one HTTP call,
one PostgREST request, one SQL transaction). The PostgREST contract
guarantees that batch insert is atomic at the row level — a single
constraint violation rolls back the whole batch.

The template builder produces a 1-sheet xlsx (Tier A) with header row +
3 sample rows so the user sees the expected shape immediately.
"""

from __future__ import annotations

import io
import logging
import time
from decimal import Decimal, InvalidOperation
from typing import Any
from uuid import UUID, uuid4

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

from api.pricing.schemas import PricingRowError, PricingUploadResponse
from api.shared.database import get_authenticated_client
from api.shared.exceptions import AppException

logger = logging.getLogger(__name__)


# Header row contract. Matches the template the frontend downloads.
EXPECTED_COLUMNS: tuple[str, ...] = ("code", "description", "unit", "price")

# In-memory cache of recent error reports so the client can download a CSV
# after a failed upload. Keyed by run_id (UUID4 string). 256-entry LRU bound
# keeps memory predictable; reports older than the bound are evicted.
_ERROR_REPORT_TTL_SECONDS = 60 * 60  # 1 hour
_ERROR_REPORT_MAX_ENTRIES = 256

_error_reports: dict[str, tuple[float, list[PricingRowError]]] = {}


def _store_error_report(errors: list[PricingRowError]) -> str:
    """Stash an error list and return its run_id.

    The frontend GETs /v1/pricing/error-report/{run_id} to download the
    CSV. Reports expire after 1 hour or when the LRU cap is hit.
    """
    # Evict expired entries first.
    now = time.monotonic()
    expired = [k for k, (ts, _) in _error_reports.items() if now - ts > _ERROR_REPORT_TTL_SECONDS]
    for k in expired:
        _error_reports.pop(k, None)

    # Evict oldest entries if over cap (drop ~25% to amortize).
    if len(_error_reports) >= _ERROR_REPORT_MAX_ENTRIES:
        ordered = sorted(_error_reports.items(), key=lambda kv: kv[1][0])
        drop = ordered[: len(ordered) // 4 or 1]
        for k, _ in drop:
            _error_reports.pop(k, None)

    run_id = str(uuid4())
    _error_reports[run_id] = (now, errors)
    return run_id


def get_error_report(run_id: str) -> list[PricingRowError] | None:
    """Look up an error report by run_id. Returns None if not found / expired."""
    entry = _error_reports.get(run_id)
    if entry is None:
        return None
    fetched_at, errors = entry
    if time.monotonic() - fetched_at > _ERROR_REPORT_TTL_SECONDS:
        _error_reports.pop(run_id, None)
        return None
    return errors


def errors_to_csv(errors: list[PricingRowError]) -> str:
    """Render a PricingRowError list as a CSV string (header + rows).

    Quoting follows RFC 4180 — wrap any field containing comma / quote /
    newline in double quotes, escape embedded quotes by doubling.
    """
    out = io.StringIO()
    out.write("row,field,message\n")
    for err in errors:
        out.write(f"{err.row},{_csv_escape(err.field or '')},{_csv_escape(err.message)}\n")
    return out.getvalue()


def _csv_escape(value: str) -> str:
    if any(c in value for c in (",", '"', "\n", "\r")):
        return '"' + value.replace('"', '""') + '"'
    return value


# ---------------------------------------------------------------------------
# Template builder
# ---------------------------------------------------------------------------


def build_template_xlsx() -> bytes:
    """Programmatically build the Tier A template workbook.

    1 sheet ('Tier A') with a header row + 3 sample rows so the user sees
    the expected shape. The frontend serves this verbatim — no signing or
    company-specific data, just a static schema-shaped starter.
    """
    wb = Workbook()
    # A fresh Workbook always exposes an active worksheet; assert so Pyright
    # narrows from `Worksheet | None`.
    ws = wb.active
    assert ws is not None
    ws.title = "Tier A"

    # Header row
    for col_idx, name in enumerate(EXPECTED_COLUMNS, start=1):
        ws.cell(row=1, column=col_idx, value=name)

    # Sample rows — chosen to mirror common Xactimate water-restoration codes
    # so the contractor sees a familiar starting point.
    samples = [
        ("WTR DRYOUT", "Water extraction & dryout (per SF)", "SF", 1.25),
        ("DRYWLL RR", 'Drywall removal & replace 1/2"', "SF", 3.50),
        ("FLR CARP RR", "Carpet pad removal & replace", "SF", 2.10),
    ]
    for row_idx, row in enumerate(samples, start=2):
        for col_idx, value in enumerate(row, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    # Reasonable column widths for readability when the user opens it.
    widths = (16, 48, 8, 10)
    for col_idx, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ---------------------------------------------------------------------------
# Upload parsing + validation
# ---------------------------------------------------------------------------


def _normalize_header(value: Any) -> str:
    """Lowercase, strip, trailing-asterisk-tolerant header normalizer.

    Excel users sometimes mark required columns with ``*`` in the header
    (e.g. ``code*``). Strip it so we still match.
    """
    if value is None:
        return ""
    s = str(value).strip().lower()
    if s.endswith("*"):
        s = s[:-1].strip()
    return s


def _coerce_price(value: Any) -> Decimal | None:
    """Best-effort numeric coercion for the price column.

    Accepts: numeric (int/float/Decimal), strings like '$1.25', '1,250.00'.
    Returns None for blank cells. Raises ValueError on anything that
    isn't parseable so the caller can attach a row-level error.
    """
    if value is None:
        return None
    if isinstance(value, Decimal):
        return value
    if isinstance(value, (int, float)):
        return Decimal(str(value))
    s = str(value).strip()
    if not s:
        return None
    # Drop currency symbols and thousands separators.
    s = s.replace("$", "").replace(",", "").strip()
    try:
        return Decimal(s)
    except InvalidOperation:
        raise ValueError(f"could not parse price: {value!r}") from None


def parse_pricing_xlsx(
    content: bytes, *, default_tier: str = "A"
) -> tuple[
    list[dict],
    list[PricingRowError],
]:
    """Parse a pricing workbook into (rows, errors).

    Returns the validated row dicts ready to insert (or empty list on any
    error) and the per-row error list. The first sheet is processed; its
    title is read as the tier. If the user named it 'Tier A' we strip the
    'Tier ' prefix; otherwise we keep the title verbatim.

    Validation:
    - header row must contain at minimum 'code' and 'price' columns
    - 'code' is required per row (string, non-empty after strip)
    - 'price' is required per row (numeric, >= 0); '$' and ',' tolerated
    - 'description' and 'unit' optional
    - rows that are entirely blank are skipped silently
    """
    errors: list[PricingRowError] = []

    try:
        wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    except Exception as e:
        # openpyxl raises various subtypes for malformed files. Surface
        # one row-level error and abort.
        return [], [PricingRowError(row=1, field=None, message=f"Cannot read workbook: {e}")]

    if not wb.sheetnames:
        return [], [PricingRowError(row=1, field=None, message="Workbook has no sheets")]

    ws = wb[wb.sheetnames[0]]

    # Tier — if sheet titled 'Tier X' use X, else use the title verbatim.
    sheet_title = (ws.title or "").strip()
    tier_name = sheet_title
    if sheet_title.lower().startswith("tier "):
        candidate = sheet_title[5:].strip()
        if candidate:
            tier_name = candidate
    if not tier_name:
        tier_name = default_tier

    rows_iter = ws.iter_rows(values_only=True)
    try:
        header = next(rows_iter)
    except StopIteration:
        return [], [PricingRowError(row=1, field=None, message="Workbook is empty")]

    header_map: dict[str, int] = {}
    for col_idx, raw in enumerate(header):
        norm = _normalize_header(raw)
        if norm:
            header_map[norm] = col_idx

    # Mandatory columns
    if "code" not in header_map:
        errors.append(
            PricingRowError(row=1, field="code", message="Missing required column 'code'")
        )
    if "price" not in header_map:
        errors.append(
            PricingRowError(row=1, field="price", message="Missing required column 'price'")
        )
    if errors:
        return [], errors

    code_idx = header_map["code"]
    price_idx = header_map["price"]
    desc_idx = header_map.get("description")
    unit_idx = header_map.get("unit")

    accepted: list[dict] = []
    seen_keys: set[tuple[str, str]] = set()

    # Spreadsheet row numbering: header is row 1, first data row is 2.
    for offset, row in enumerate(rows_iter, start=2):
        # Skip entirely empty rows silently.
        if row is None:
            continue
        if all(cell is None or (isinstance(cell, str) and not cell.strip()) for cell in row):
            continue

        row_errors: list[PricingRowError] = []

        code_val = row[code_idx] if code_idx < len(row) else None
        price_val = row[price_idx] if price_idx < len(row) else None

        code_str = "" if code_val is None else str(code_val).strip()
        if not code_str:
            row_errors.append(PricingRowError(row=offset, field="code", message="code is required"))

        try:
            price_dec = _coerce_price(price_val)
        except ValueError as e:
            row_errors.append(PricingRowError(row=offset, field="price", message=str(e)))
            price_dec = None

        if price_dec is None:
            already_flagged = any(err.field == "price" for err in row_errors)
            if not already_flagged:
                row_errors.append(
                    PricingRowError(row=offset, field="price", message="price is required")
                )
        elif price_dec < 0:
            row_errors.append(
                PricingRowError(row=offset, field="price", message="price must be >= 0")
            )

        description = None
        if desc_idx is not None and desc_idx < len(row):
            d = row[desc_idx]
            description = None if d is None else str(d).strip() or None

        unit = None
        if unit_idx is not None and unit_idx < len(row):
            u = row[unit_idx]
            unit = None if u is None else str(u).strip() or None

        if row_errors:
            errors.extend(row_errors)
            continue

        # If we reached here without row_errors, price_dec was successfully
        # coerced AND non-negative (the None-branch above appends an error
        # which would have made us continue). Narrow the type for Pyright.
        assert price_dec is not None

        # In-file dedup against (code, tier) so we don't surface a
        # confusing 23505 from the DB after a successful "looking" upload.
        key = (code_str, tier_name)
        if key in seen_keys:
            errors.append(
                PricingRowError(
                    row=offset,
                    field="code",
                    message=f"Duplicate code '{code_str}' for tier '{tier_name}' in this file",
                )
            )
            continue
        seen_keys.add(key)

        accepted.append(
            {
                "code": code_str,
                "description": description,
                "unit": unit,
                "price": float(price_dec),
                "tier": tier_name,
            }
        )

    if errors:
        return [], errors

    return accepted, []


async def persist_pricing_rows(
    token: str,
    company_id: UUID,
    rows: list[dict],
) -> int:
    """Insert validated pricing rows for the user's company.

    Uses upsert on (company_id, code, tier) so re-uploading the same file
    is idempotent (Pre-launch decision: re-upload should refresh prices,
    not error out). One PostgREST call ⇒ one SQL transaction.
    """
    if not rows:
        return 0

    client = await get_authenticated_client(token)

    payload = [{"company_id": str(company_id), **r} for r in rows]

    try:
        result = await (
            client.table("scope_codes")
            .upsert(payload, on_conflict="company_id,code,tier")
            .execute()
        )
    except Exception as e:
        logger.error("pricing upsert failed for company %s: %s", company_id, e)
        raise AppException(
            status_code=500,
            detail=f"Failed to persist pricing rows: {e}",
            error_code="PRICING_PERSIST_FAILED",
        )

    return len(result.data or [])


async def upload_pricing_file(
    token: str,
    company_id: UUID,
    content: bytes,
) -> PricingUploadResponse:
    """End-to-end pricing upload: parse → validate → persist on success.

    On validation failure, nothing is persisted and we register an error
    report keyed by run_id for CSV download.
    """
    rows, errors = parse_pricing_xlsx(content)
    if errors:
        run_id = _store_error_report(errors)
        # Tier might not be reliable on a malformed file; surface 'unknown'
        # rather than fabricate. The UI only uses tier on the success path.
        return PricingUploadResponse(
            items_loaded=0,
            tier="unknown",
            errors=errors,
            run_id=run_id,
        )

    items_loaded = await persist_pricing_rows(token, company_id, rows)
    tier = rows[0]["tier"] if rows else "A"
    return PricingUploadResponse(
        items_loaded=items_loaded,
        tier=tier,
        errors=[],
        run_id=None,
    )
